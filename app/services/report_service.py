"""
app/services/report_service.py

Report generation in PDF / CSV / Excel using pandas, ReportLab and OpenPyXL.

Reports are built from the shared database via the *sync* SQLAlchemy engine
(pandas.read_sql) and written to ``settings.REPORTS_DIR``. Metadata for each
generated file is tracked in the ``analytics_report_jobs`` table so the
``/reports/{id}/download`` endpoint can stream it back later.
"""
import asyncio
import json
import logging
import os
from datetime import date, datetime, timezone
from typing import Optional

import pandas as pd
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
    Paragraph,
)
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.database import sync_engine
from app.models.events import ReportJob

logger = logging.getLogger("analytics.reports")

VALID_TYPES = {"sales", "inventory", "customers"}
VALID_FORMATS = {"pdf", "csv", "xlsx"}


def _status_in() -> str:
    return ", ".join(f"'{s}'" for s in settings.SALE_STATUSES)


# ── Dataset builders (sync, pandas) ────────────────────────────
def _sales_dataframe(start: Optional[date], end: Optional[date]) -> pd.DataFrame:
    sql = f"""
        SELECT
            b.title                              AS book,
            b.author                             AS author,
            SUM(oi.quantity)                     AS units_sold,
            ROUND(SUM(oi.quantity * oi.unit_price), 2) AS revenue
        FROM order_items oi
        JOIN orders o ON o.id = oi.order_id
        JOIN books b  ON b.id = oi.book_id
        WHERE o.status IN ({_status_in()})
          AND (CAST(:start AS date) IS NULL OR o.created_at >= CAST(:start AS date))
          AND (CAST(:end AS date) IS NULL OR o.created_at < (CAST(:end AS date) + INTERVAL '1 day'))
        GROUP BY b.title, b.author
        ORDER BY revenue DESC
    """
    with sync_engine.connect() as conn:
        return pd.read_sql(text(sql), conn, params={"start": start, "end": end})


def _inventory_dataframe(*_args) -> pd.DataFrame:
    sql = """
        SELECT
            b.title                                AS book,
            COALESCE(i.quantity, b.stock, 0)       AS stock,
            COALESCE(i.reorder_level, 10)          AS reorder_level,
            b.price                                AS price,
            ROUND(COALESCE(i.quantity, b.stock, 0) * b.price, 2) AS stock_value
        FROM books b
        LEFT JOIN inventory_items i ON i.book_id = b.id
        WHERE b.is_active = TRUE
        ORDER BY stock ASC
    """
    with sync_engine.connect() as conn:
        return pd.read_sql(text(sql), conn)


def _customers_dataframe(*_args) -> pd.DataFrame:
    sql = f"""
        SELECT
            u.email                          AS email,
            COUNT(o.id)                      AS orders,
            ROUND(COALESCE(SUM(o.total_amount), 0), 2) AS lifetime_value
        FROM users u
        LEFT JOIN orders o
               ON o.user_id = u.id AND o.status IN ({_status_in()})
        GROUP BY u.email
        ORDER BY lifetime_value DESC
    """
    with sync_engine.connect() as conn:
        return pd.read_sql(text(sql), conn)


_BUILDERS = {
    "sales": _sales_dataframe,
    "inventory": _inventory_dataframe,
    "customers": _customers_dataframe,
}


def build_dataset_records(report_type: str, start=None, end=None) -> list[dict]:
    """Build a report dataset and return it as a list of plain dict rows."""
    report_type = report_type.lower()
    if report_type not in _BUILDERS:
        raise ValueError(f"Unsupported report_type '{report_type}'.")
    df = _BUILDERS[report_type](start, end)
    # Normalise numpy/pandas scalar types to native Python for JSON.
    return df.to_dict(orient="records")


# ── File writers (sync) ────────────────────────────────────────
def _write_csv(df: pd.DataFrame, path: str) -> None:
    df.to_csv(path, index=False)


def _write_xlsx(df: pd.DataFrame, path: str, sheet: str) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet[:31])
        ws = writer.sheets[sheet[:31]]
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        # Auto-fit column widths.
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(width + 2, 50)


def _write_pdf(df: pd.DataFrame, path: str, title: str) -> None:
    doc = SimpleDocTemplate(path, pagesize=A4, title=title)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"]), Spacer(1, 12)]
    elements.append(
        Paragraph(
            f"Generated {datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC} — "
            f"{len(df)} rows",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 12))

    # Cap rows so PDFs stay reasonable.
    display = df.head(200)
    data = [list(display.columns)] + display.astype(str).values.tolist()
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EFF3F8")]),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)


def _generate_file(report_type: str, file_format: str, start, end) -> tuple[str, str]:
    """Build the dataset and write the file. Returns (file_path, file_name)."""
    os.makedirs(settings.REPORTS_DIR, exist_ok=True)
    df = _BUILDERS[report_type](start, end)

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    file_name = f"{report_type}_report_{stamp}.{file_format}"
    file_path = os.path.join(settings.REPORTS_DIR, file_name)
    title = f"{report_type.capitalize()} Report"

    if file_format == "csv":
        _write_csv(df, file_path)
    elif file_format == "xlsx":
        _write_xlsx(df, file_path, report_type)
    else:  # pdf
        _write_pdf(df, file_path, title)
    return file_path, file_name


# ── Async orchestration ────────────────────────────────────────
class ReportService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def generate(
        self,
        report_type: str,
        file_format: str,
        start: Optional[date] = None,
        end: Optional[date] = None,
    ) -> ReportJob:
        report_type = report_type.lower()
        file_format = file_format.lower()
        if report_type not in VALID_TYPES:
            raise ValueError(f"Unsupported report_type '{report_type}'.")
        if file_format not in VALID_FORMATS:
            raise ValueError(f"Unsupported file_format '{file_format}'.")

        job = ReportJob(
            report_type=report_type,
            file_format=file_format,
            status="pending",
            params=json.dumps(
                {"start": str(start) if start else None, "end": str(end) if end else None}
            ),
        )
        self.db.add(job)
        await self.db.commit()
        await self.db.refresh(job)

        try:
            # Run the blocking pandas/reportlab work off the event loop.
            file_path, file_name = await asyncio.to_thread(
                _generate_file, report_type, file_format, start, end
            )
            job.file_path = file_path
            job.file_name = file_name
            job.status = "ready"
            job.completed_at = datetime.now(timezone.utc)
        except Exception as exc:  # pragma: no cover - defensive
            logger.exception("Report generation failed")
            job.status = "failed"
            job.error = str(exc)
        await self.db.commit()
        await self.db.refresh(job)
        return job

    async def get(self, report_id: str) -> Optional[ReportJob]:
        return await self.db.get(ReportJob, report_id)
